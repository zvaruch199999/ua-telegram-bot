from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

FILE = "data/offers.xlsx"

HEADERS = [
    "ID", "Дата створення", "Категорія", "Тип", "Вулиця", "Місто", "Район",
    "Переваги", "Оренда", "Депозит", "Комісія", "Паркінг",
    "Заселення", "Огляди", "Маклер", "Статус",
    "Хто знайшов нерухомість", "Хто знайшов клієнта",
    "Дата контракту", "Сума провізії", "Оплати",
    "Клієнт", "ПМЖ", "Контакт"
]

def init_excel():
    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(FILE)

def add_offer(data: dict):
    wb = load_workbook(FILE)
    ws = wb.active
    row_id = ws.max_row
    ws.append([
        row_id,
        datetime.now().strftime("%Y-%m-%d"),
        data["category"],
        data["type"],
        data["street"],
        data["city"],
        data["district"],
        data["advantages"],
        data["rent"],
        data["deposit"],
        data["commission"],
        data["parking"],
        data["move_in"],
        data["viewing"],
        data["broker"],
        "Активна"
    ])
    wb.save(FILE)
    return row_id
